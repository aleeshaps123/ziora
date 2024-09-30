from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
from .forms import ProductForm
from .models import Product, UserProfile, Order
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.contrib.auth.hashers import make_password
from django.contrib import messages
from .forms import UserProfileForm
from django.urls import reverse, reverse_lazy
from .models import Cart, Product
from django.contrib.auth import authenticate, login ,logout
from django.db.models import F
from django.contrib.auth.decorators import login_required
from .models import *




from django.shortcuts import render
from django.http import JsonResponse
import requests

# Replace with your API key and URL
API_KEY = 'AIzaSyD4LfL36lQf1XehsiYjYC14NEC5d7xd4NI'
API_URL = 'https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash-latest:generateContent'

# Define predefined responses for the chatbot
PREDEFINED_RESPONSES = {
    # Add your predefined chatbot responses here (same as before)
    # ...
}

def index(request):
    if request.method == 'POST':
        # Chatbot interaction
        user_message = request.POST.get('message')
        conversation_history = request.session.get('conversation_history', [])

        conversation_history.append(f"input: {user_message}")
        bot_reply = PREDEFINED_RESPONSES.get(user_message, None)

        if not bot_reply:
            # Make a request to the Google API if the message is not predefined
            headers = {'Content-Type': 'application/json'}
            messages = [{'text': message} for message in conversation_history]

            data = {'contents': [{'parts': messages}]}

            try:
                response = requests.post(f'{API_URL}?key={API_KEY}', headers=headers, json=data)
                response.raise_for_status()
                api_response = response.json()
                bot_reply = api_response['candidates'][0]['content']['parts'][0]['text']
                bot_reply = '. '.join(bot_reply.split('. ')[:3])  # Limit response to 3 sentences
            except requests.RequestException as e:
                print(f"API request error: {e}")
                bot_reply = 'Sorry, there was an error processing your request.'

        conversation_history.append(f"output: {bot_reply}")
        request.session['conversation_history'] = conversation_history
        return JsonResponse({'reply': bot_reply})

    # Existing index logic for products and categories
    categories = Product.objects.values_list('category', flat=True).distinct()
    selected_category = request.GET.get('category')

    if selected_category:
        products = Product.objects.filter(category=selected_category)
    else:
        products = Product.objects.all()

    in_stock_products = products.exclude(quantity=0).order_by('-created_at')[:6]
    out_of_stock_products = products.filter(quantity=0)

    context = {
        'categories': categories,
        'in_stock_products': in_stock_products,
        'out_of_stock_products': out_of_stock_products
    }

    return render(request, 'index.html', context)


def products(request):
    # Fetch distinct categories for filtering
    categories = Product.objects.values_list('category', flat=True).distinct()
    
    selected_category = request.GET.get('category')
    if selected_category:
        # Filter products by selected category
        products = Product.objects.filter(category=selected_category)
    else:
        # Fetch all products
        products = Product.objects.all()
    
    # Filter out products with zero quantity
    in_stock_products = products.exclude(quantity=0).order_by('-created_at')[:]
    out_of_stock_products = products.filter(quantity=0)
    
    context = {
        'categories': categories,
        'in_stock_products': in_stock_products,
        'out_of_stock_products': out_of_stock_products
    }
    return render(request, 'product.html', context)


from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib.auth.forms import AuthenticationForm

def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                if user.is_superuser:
                    return redirect('adminpage')  # Redirect admin to admin page
                else:
                    return redirect('index')  # Redirect regular user to index page
            else:
                error_message = 'Invalid credentials. Please try again.'
                return render(request, 'user_login.html', {'form': form, 'error_message': error_message})
    else:
        form = AuthenticationForm()
    
    return render(request, 'user_login.html', {'form': form})


from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.core.mail import send_mail
from .forms import CustomUserCreationForm

def user_register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Log the user in after registration
            login(request, user)

            # Send registration email
            subject = 'Welcome to Ziora Collections'
            message = f'Hi {user.username},\n\nThank you for registering on Ziora!\n\nYour username: {user.username}\nYour password: {form.cleaned_data["password1"]}'
            from_email = 'aleeshasibi26@gmail.com'
            to_email = user.email
            send_mail(subject, message, from_email, [to_email])

            return redirect('user_login')  # Redirect to login page after registration
    else:
        form = CustomUserCreationForm()
    return render(request, 'user_register.html', {'form': form})




from django.contrib.auth.forms import PasswordChangeForm
@login_required
def profile_update(request):
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=request.user)
        password_form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid() and password_form.is_valid():
            user = form.save(commit=False)
            password = password_form.cleaned_data['new_password1']
            if password:
                user.password = make_password(password)  # Change password
            user.save()
            password_form.save()
            messages.success(request, 'Your profile has been updated successfully.')
            return redirect('user_login')  # Redirect to the profile update page
    else:
        form = UserProfileForm(instance=request.user)
        password_form = PasswordChangeForm(request.user)
    return render(request, 'profile_update.html', {'form': form, 'password_form': password_form})

def cart(request):
    upcart_items = Cart.objects.filter(user=request.user)
    
    
    # <QuerySet [<Cart: Cart for alvin: Apple VX max>, <Cart: Cart for alvin: Apple VX max>]>
    cart_items={}
    for item in upcart_items:
        # print(item.product.name)
        
        if item.product.name in cart_items:
            cart_items[item.product.name]['quantity']+=1
            cart_items[item.product.name]['total_price']+=cart_items[item.product.name]['price']
            
        else:
            cart_items[item.product.name]={
                'id':item.product.id,
                'name':item.product.name,
                'price':item.product.price,
                'quantity':item.quantity,
                'price':item.product.price,
                'total_price':item.product.price,
                'image':item.product.image
            }
            # print(cart_items)

    total_price = sum(item['total_price'] for item in cart_items.values())

    
    context= {
        'cart_items': cart_items,
        'total_price': total_price
        }
    
    return render(request, 'cart.html',context)

from django.core.mail import send_mail

def checkout_view(request):
    cart_items = Cart.objects.filter(user=request.user)
    total_amount = sum(item.quantity * item.product.price for item in cart_items)

    if request.method == 'POST':
        fullname = request.POST.get('fullname')
        address = request.POST.get('address')
        phone_number = request.POST.get('phone_number')
        city = request.POST.get('city')
        postal_code = request.POST.get('postal_code')

        # Construct the redirect URL for order summary
        redirect_url = reverse('order_summary')
        params = {
            'fullname': fullname,
            'address': address,
            'phone_number': phone_number,
            'city': city,
            'postal_code': postal_code,
            'total_amount': total_amount,
        }
        redirect_url += '?' + '&'.join([f"{key}={value}" for key, value in params.items()])

        # Send email confirmation to the logged-in user's email address
        subject = 'Order Confirmation - ZIORA COLLECTIONS'
        message = f"Dear {fullname},\n\n"
        message += f"Thanks for ordering with Ziora.\n\n"
        message += f"For order details, click the following link:\n"
        message += f"Or Please click TRACK ORDER button in your Ziora account "
        message += "Best regards,\nZiora Team"

        sender_email = 'aleeshasibi26@gmail.com'  # Replace with your sender email
        recipient_email = request.user.email  # Use the logged-in user's email address
        send_mail(subject, message, sender_email, [recipient_email], html_message=message)

        # Redirect to order summary page
        return HttpResponseRedirect(redirect_url)

    return render(request, 'checkout.html', {'total_amount': total_amount})


from .models import Cart
from django.contrib.auth import get_user_model

User = get_user_model()

@login_required(login_url=user_login)
def add_to_cart(request, product_id):

    product = Product.objects.get(pk=product_id)
    
    if request.method == 'POST':
        if product.quantity > 0:
            cart_item = Cart(user=request.user, product=product, quantity=1)
            cart_item.save()
            product.quantity -= 1
            product.save()
            return redirect('cart')
        else:
            return render(request, 'out_of_stock.html', {'product': product})
    
    # Add this return statement for the GET request method
    return redirect('cart')

    

def decrease_to_cart(request, product_id):
    cart_items = Cart.objects.filter(product_id=product_id,user=request.user).first()
    print('============')
    print("items",cart_items)

    cart_items.delete()
    return redirect('cart')



def remove_from_cart(request, product_id):
    print('--------')
    cart_items = Cart.objects.filter(product_id=product_id,user=request.user).all()
    print('============')
    print("items",cart_items)
    cart_items.delete()
    return redirect('cart')

@login_required(login_url=user_login)
def cart(request):
    upcart_items = Cart.objects.filter(user=request.user)
    
    
    # <QuerySet [<Cart: Cart for alvin: Apple VX max>, <Cart: Cart for alvin: Apple VX max>]>
    cart_items={}
    for item in upcart_items:
        # print(item.product.name)
        
        if item.product.name in cart_items:
            cart_items[item.product.name]['quantity']+=1
            cart_items[item.product.name]['total_price']+=cart_items[item.product.name]['price']
            
        else:
            cart_items[item.product.name]={
                'id':item.product.id,
                'name':item.product.name,
                'price':item.product.price,
                'quantity':item.quantity,
                'price':item.product.price,
                'total_price':item.product.price,
                'image':item.product.image
            }
            # print(cart_items)

    total_price = sum(item['total_price'] for item in cart_items.values())

    
    context= {
        'cart_items': cart_items,
        'total_price': total_price
        }
    
    return render(request, 'cart.html',context)


def admin_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None and user.is_superuser:
                login(request, user)
                url='/adminpage'
                x=f'''
                    <script>
                        alert("wlecome admin");
                        window.location.href = "{url}"; 
                    </script>
                '''
                
                return HttpResponse(x)
        else:
            form = AuthenticationForm()  
            error_message = 'Invalid credentials. Please try again.'
            return render(request, 'admin_login.html', {'form': form, 'error_message': error_message})
    
    else:
        form = AuthenticationForm()

    return render(request, 'admin_login.html', {'form': form})

def adminpage(request):
    products_at_reorder_level = Product.objects.filter(quantity__lte=F('reorderlevel'))
    products = Product.objects.all()
    categories = Product.objects.values_list('category', flat=True).distinct()
    categories = Product.objects.values_list('category', flat=True).distinct()
    
    selected_category = request.GET.get('category')
    if selected_category:
        products = Product.objects.filter(category=selected_category)
    else:
        products = Product.objects.all()

    context = {
        'products_at_reorder_level': products_at_reorder_level,
        'productdata': products,
        'categories': categories, 'products': products
    }
    
    return render(request, 'admin.html', context)

    


def admin_add(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            # Get the form data including quantity and reorderlevel
            name = form.cleaned_data['name']
            description = form.cleaned_data['description']
            price = form.cleaned_data['price']
            image = form.cleaned_data['image']
            category = form.cleaned_data['category']
            quantity = form.cleaned_data['quantity']
            reorderlevel = form.cleaned_data['reorderlevel']

            # Create a new product instance
            new_product = Product.objects.create(
                name=name,
                description=description,
                price=price,
                image=image,
                category=category,
                quantity=quantity,
                reorderlevel=reorderlevel
            )
            # Save the new product to the database
            new_product.save()

            return redirect('adminpage')  # Redirect after adding a product
    else:
        form = ProductForm()

    return render(request, 'admin_add.html', {'form': form})

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import Order

@login_required
def admin_order_view(request):
    # Get the status filter from the GET parameters (if any)
    selected_status = request.GET.get('status', 'all')

    # Filter orders based on the selected status
    if selected_status == 'all':
        orders = Order.objects.all()
    else:
        orders = Order.objects.filter(status=selected_status)

    # Pass the selected status to the template to maintain the selected option in the dropdown
    context = {
        'orders': orders,
        'selected_status': selected_status,
    }

    return render(request, 'order_list_and_detail.html', context)

@login_required
def user_order_view(request):
    # user view
    orders = Order.objects.all()
    context = {'orders': orders}
    return render(request, 'order_list_and_detail_user.html', context)

@login_required
def order_detail_view(request, order_id):
    # Admin view to display order details
    order = Order.objects.get(id=order_id)
    items = order.items.all()
    context = {'order': order, 'items': items}
    return render(request, 'order_detail.html', context)

def user_logout(request):
    logout(request)
    # Redirect to the index page after logout

    return redirect('index')

def edit_product(request, product_id):
    product_to_edit = get_object_or_404(Product, pk=product_id)

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product_to_edit)
        if form.is_valid():
            # Check if any of the fields have negative values
            price = form.cleaned_data['price']
            quantity = form.cleaned_data['quantity']
            reorderlevel = form.cleaned_data['reorderlevel']
            
            if price >= 0 and quantity >= 0 and reorderlevel >= 0:
                form.save()
                return redirect('adminpage')  # Redirect to the admin page if all values are non-negative
            else:
                # If any field has a negative value, display an error message
                form.add_error(None, "Price, quantity, and reorder level cannot be negative.")
    else:
        form = ProductForm(instance=product_to_edit)

    return render(request, 'edit_product.html', {'form': form, 'product': product_to_edit})


def delete_product(request, product_id):
    product_to_delete = get_object_or_404(Product, pk=product_id)  # Assuming product is the name of model

    if request.method == 'POST':
        confirmation = request.POST.get('confirmation', None)
        if confirmation == 'confirmed':
            #product_to_delete.is_active = False  # Set the product as inactive
            product_to_delete.delete()
            return redirect('adminpage')  # Redirect to the user page after deletion
        else:
            return redirect('adminpage')  # Redirect without deleting if not confirmed

    return render(request, 'confirmation.html', {'product': product_to_delete})
from django.core.paginator import Paginator
from django.shortcuts import render
from .models import Order

def order_list_and_detail(request):
    # Ensure the user is authenticated
    if not request.user.is_authenticated:
        return redirect('login')  # Redirect to login if not authenticated

    # Get the logged-in user
    user = request.user
    
    # Get the status filter from the request
    status_filter = request.GET.get('status', 'all')

    # Fetch orders based on user and filter
    if status_filter == 'all':
        orders = Order.objects.filter(user=user)  # Filter by user
    else:
        orders = Order.objects.filter(user=user, status=status_filter)

    # Set up pagination
    paginator = Paginator(orders, 10)  # Show 10 orders per page
    page_number = request.GET.get('page')
    orders_page = paginator.get_page(page_number)

    context = {
        'orders': orders_page,
        'selected_status': status_filter,
    }
    return render(request, 'order_list_and_detail.html', context)
@login_required
def user_order_view(request):
    # Admin view to display all orders
    orders = Order.objects.filter(user=request.user).all()
    context = {'orders': orders}
    return render(request, 'order_list_and_detail.html', context)




from django.shortcuts import render, redirect
from .models import Cart, Order

from collections import defaultdict
from django.utils import timezone

def order_summary_view(request):
    # Retrieve order information from request parameters
    fullname = request.GET.get('fullname')
    address = request.GET.get('address')
    phone_number = request.GET.get('phone_number')
    city = request.GET.get('city')
    postal_code = request.GET.get('postal_code')
    total_amount = request.GET.get('total_amount')



    # Get the current date
    order_date = timezone.now().date()

    # Create order instance
    order = Order.objects.create(
        user=request.user,
        fullname=fullname,
        address=address,
        city=city,
        postal_code=postal_code,
        total_amount=total_amount,
    )

    # Retrieve and move cart items to the order
    cart_items = Cart.objects.filter(user=request.user)
    product_quantities = defaultdict(int)  # To store total quantities of each product

    for cart_item in cart_items:
        order.items.create(product=cart_item.product, quantity=cart_item.quantity)
        product_quantities[cart_item.product] += cart_item.quantity
        cart_item.delete()

    # Prepare context for rendering the order summary page
    product_details = []  # To store aggregated product details
    for product, quantity in product_quantities.items():
        product_details.append({'product': product, 'quantity': quantity})

    context = {
        'fullname': fullname,
        'address': address,
        'phone_number':phone_number,
        'city': city,
        'postal_code': postal_code,
        'total_amount': total_amount,
        'product_details': product_details,
        'order_date': order_date,  # Pass the current order date to the template
        'user':request.user.email,
    }

    return render(request, 'order_summary.html', context)



# views.py
from django.shortcuts import render, get_object_or_404, redirect
from .models import Product, Review
from .forms import ReviewForm
from django.contrib.auth.decorators import login_required

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import Product, Review
from .forms import ReviewForm  # Make sure you have this form created properly
def product_detail(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    reviews = product.reviews.all()

    if request.method == 'POST':
        if request.user.is_authenticated:
            form = ReviewForm(request.POST)
            if form.is_valid():
                review = form.save(commit=False)
                review.product = product
                review.user = request.user
                review.save()
                return redirect('product_detail', product_id=product_id)
        else:
            return redirect('login')  # Redirect to login if user is not authenticated
    else:
        form = ReviewForm()

    # Create a range for star ratings
    star_range = range(1, 6)

    context = {
        'product': product,
        'reviews': reviews,
        'form': form,
        'star_range': star_range
    }
    return render(request, 'product_detail.html', context)

def product_list(request): #admin page product view
    products = Product.objects.all()
    return render(request, 'user/admin.html', {'product': products})

def pop_message(url,message):
    url=url
    x=f'''
        <script>
            alert("{message}");
            window.location.href = "{url}"; 
        </script>
    '''
    return(x)

def search(request):
    query = request.GET.get('query')
    products = Product.objects.filter(name__icontains=query, is_active=True)
    return render(request, 'search_results.html', {'products': products, 'query': query})

def search_view(request):
    query = request.GET.get('query', '')  # Get the 'query' parameter from the request's GET parameters

    if query:
        # Perform a search using the query, ensuring it's not None
        products = Product.objects.filter(name__icontains=query)
    else:
        # Handle the case where 'query' is None or empty
        products = []

    context = {
        'query': query,
        'products': products,
    }

    return render(request, 'search_results.html', context)


from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404

@login_required
def update_status(request, order_id):
    if request.method == 'POST':
        new_status = request.POST.get('status')
        order = get_object_or_404(Order, pk=order_id)
        order.status = new_status
        order.save()
        
        return JsonResponse({
            'success': True,
            'new_status_display': order.get_status_display()
        })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

def stock(request):
    products_at_reorder_level = Product.objects.filter(quantity__lte=F('reorderlevel'))
    products = Product.objects.all()

    context = {
        'products_at_reorder_level': products_at_reorder_level,
        'productdata': products
    }
    return render(request,'stock.html',context)

def about_us(request):
    return render(request, 'about.html')




def product_list(request): #admin page product view
    products = Product.objects.all()
    return render(request, 'saree/admin.html', {'product': products})

@login_required(login_url=user_login)
def add_to_wishlist(request, product_id):
    if 'wishlist' not in request.session:  
        request.session['wishlist'] = []  

    wishlist = request.session['wishlist'] 
    if product_id not in wishlist:  
        wishlist.append(product_id)
        request.session['wishlist'] = wishlist 
        message='Item added to your wishlist!'
    else:
        message='Item already in wishlist'

    url='/'
    return HttpResponse(pop_message(url,message))

@login_required(login_url=user_login)
def remove_to_wishlist(request, product_id):
    wishlist = request.session['wishlist']   
    wishlist.remove(product_id)
    request.session['wishlist']=wishlist
    return redirect('index')

@login_required(login_url=user_login)
def view_to_wishlist(request):
    if 'wishlist' not in request.session:  
        context={
            'product':'',
        }
    else:
        wishlist = request.session['wishlist'] 
        data=Product.objects.filter(id__in=wishlist)
        
        context={
            'data':data
        }

    return render(request,'wishlist.html',context)


from django.shortcuts import render, get_object_or_404, redirect
from .models import Supplier, SupplierContract, CommunicationHistory
from .forms import SupplierForm, SupplierContractForm, CommunicationHistoryForm

def add_supplier(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save(commit=False)
            username = form.cleaned_data.get('email')
            password = generate_random_password()
            user = UserProfile.objects.create_user(username=username,email=supplier.email,password=password)
            supplier.user = user
            supplier.save()

            send_mail(
                'supplier account created',
                f'your supplier account has been created.\nUsername:{username} \nPassword: {password}',
                'aleeshasibi26@gmail.com',
                [supplier.email],
                fail_silently=False,
            )
            
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    return render(request, 'add_supplier.html', {'form': form})

def edit_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    return render(request, 'edit_supplier.html', {'form': form})

def delete_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    supplier.delete()
    return redirect('supplier_list')

def supplier_list(request):
    suppliers = Supplier.objects.all()
    return render(request, 'supplier_list.html', {'suppliers': suppliers})

def supplier_login(request):
    if request.method == 'POST':
        form= AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None and hasattr(user,'supplier_profile'):
                login(request,user)
                return redirect ('supplier_dashboard')
    else:
        form = AuthenticationForm()
    return render(request,'supplier_login.html',{'form':form})


def supplier_dashboard(request):
    return render(request,'supplier_dashboard.html')

from django.core.mail import send_mail
import random
import string


def generate_random_password(length=8):
    characters=string.ascii_letters + string.digits
    return ''.join(random.choice(characters)for i in range(length))

from django.shortcuts import render, redirect
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.views import PasswordResetView
from django.contrib import messages


class CustomPasswordResetView(PasswordResetView):
    template_name = 'reset_password.html'
    form_class = PasswordResetForm

    def form_valid(self, form):
        messages.success(self.request, "Password reset email has been sent.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Password reset email could not be sent. Please try again.")
        return super().form_invalid(form)





from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from .forms import SellerRegistrationForm, SellerForm
from .models import Seller

def register_seller(request):
    if request.method == 'POST':
        user_form = SellerRegistrationForm(request.POST)
        seller_form = SellerForm(request.POST)
        if user_form.is_valid() and seller_form.is_valid():
            user = user_form.save(commit=False)
            user.set_password(user_form.cleaned_data['password'])
            user.save()
            seller = seller_form.save(commit=False)
            seller.user = user
            seller.save()
            return redirect('login')
    else:
        user_form = SellerRegistrationForm()
        seller_form = SellerForm()
    return render(request, 'register_seller.html', {
        'user_form': user_form,
        'seller_form': seller_form
    })

def login_seller(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('seller_dashboard')
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})

@login_required
def seller_dashboard(request):
    seller = Seller.objects.get(user=request.user)
    return render(request, 'seller_dashboard.html', {'seller': seller})



from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .models import Seller

@login_required
def seller_list(request):
    sellers = Seller.objects.all()
    return render(request, 'admin/seller_list.html', {'sellers': sellers})

@login_required
def delete_seller(request, pk):
    seller = get_object_or_404(Seller, pk=pk)
    user = seller.user
    seller.delete()
    user.delete()
    return redirect('seller_list')

@login_required
def edit_seller(request, pk):
    seller = get_object_or_404(Seller, pk=pk)
    if request.method == 'POST':
        form = SellerForm(request.POST, instance=seller)
        if form.is_valid():
            form.save()
            return redirect('seller_list')
    else:
        form = SellerForm(instance=seller)
    return render(request, 'admin/edit_seller.html', {'form': form})


from django.shortcuts import render, redirect
from .models import Product, PurchaseOrder
from .forms import PurchaseOrderForm

def create_purchase_order(request, product_id):
    product = Product.objects.get(id=product_id)
    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST)
        if form.is_valid():
            purchase_order = form.save(commit=False)
            purchase_order.product = product  # Set the product for the purchase order
            purchase_order.save()

            
            return redirect('stock')  # Redirect to stock page after creating purchase order
    else:
        form = PurchaseOrderForm(initial={'product': product})
    return render(request, 'purchase_order.html', {'form': form, 'product': product})


def purchase_order_list(request):
    purchase_orders = PurchaseOrder.objects.all()
    return render(request, 'purchase_order_list.html', {'purchase_orders': purchase_orders})

@login_required
def seller_dashboard(request):
    # Get the current seller
    current_seller = request.user.seller

    # Query purchase orders associated with the current seller
    purchase_orders = PurchaseOrder.objects.filter(seller=current_seller)

    return render(request, 'seller_dashboard.html', {'purchase_orders': purchase_orders})

from .forms import PurchaseOrderUpdateForm

def update_purchase_order(request, order_id):
    order = get_object_or_404(PurchaseOrder, id=order_id)
    if request.method == 'POST':
        form = PurchaseOrderUpdateForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            return redirect('seller_dashboard')  # Redirect to seller dashboard after updating
    else:
        form = PurchaseOrderUpdateForm(instance=order)
    return render(request, 'update_purchase_order.html', {'form': form, 'order': order})


from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Product, Review  # Import your Review model

def delete_review(request, review_id):
    review = get_object_or_404(Review, id=review_id)
    if request.method == 'POST':
        review.delete()
        messages.success(request, 'Review deleted successfully.')
    return redirect('product_detail', product_id=review.product.id)  # Redirect back to the product detail page


from django.shortcuts import render, get_object_or_404
from .models import Product, Review

def product_comments(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    reviews = Review.objects.filter(product=product) # Adjust as needed based on your model structure
    star_range=range(1,6)
    context = {
        'product': product,
        'reviews': reviews,
        'star_range':star_range
    }
    return render(request, 'product_comments.html', context)

from django.shortcuts import render
from .models import Product, Review
from django.db.models import Count, Case, When, IntegerField

@login_required(login_url=user_login)
def sentiment_analysis(request):
    products = Product.objects.all()
    selected_product_id = request.GET.get('product')
    sentiment_data = {}
    positive_comments = []
    negative_comments = []
    neutral_comments = []

    if selected_product_id:
        product = Product.objects.get(id=selected_product_id)
        reviews = Review.objects.filter(product=product)

        total_reviews = reviews.count()
        if total_reviews > 0:
            sentiment_counts = reviews.aggregate(
                positive=Count(Case(When(rating__gte=4, then=1), output_field=IntegerField())),
                negative=Count(Case(When(rating__lte=2, then=1), output_field=IntegerField())),
                neutral=Count(Case(When(rating__gt=2, rating__lt=4, then=1), output_field=IntegerField()))
            )

            positive_count = sentiment_counts['positive']
            negative_count = sentiment_counts['negative']
            neutral_count = sentiment_counts['neutral']

            sentiment_data = {
                'Positive': positive_count,
                'Negative': negative_count,
                'Neutral': neutral_count,
            }

            # Retrieve comments for each sentiment
            positive_comments = reviews.filter(rating__gte=4)
            negative_comments = reviews.filter(rating__lte=2)
            neutral_comments = reviews.filter(rating__gt=2, rating__lt=4)

    context = {
        'products': products,
        'selected_product_id': selected_product_id,
        'sentiment_data': sentiment_data,
        'positive_comments': positive_comments,
        'negative_comments': negative_comments,
        'neutral_comments': neutral_comments,
    }

    return render(request, 'sentiment.html', context)


# views.py

from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, F
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from .models import Order, OrderItem  # Adjust based on your actual models
@login_required(login_url=user_login)
def monthly_sales_report(request):
    # Get the month from the query parameters, default to the current month and year
    month = request.GET.get('month', datetime.now().strftime('%Y-%m'))
    
    # Split the month and year for query
    year, month = map(int, month.split('-'))

    # Calculate the start and end dates for the selected month
    start_date = datetime(year, month, 1)
    if month < 12:
        end_date = datetime(year, month + 1, 1)
    else:
        end_date = datetime(year + 1, 1, 1)

    # Filter orders within the selected month
    orders = Order.objects.filter(created_at__range=[start_date, end_date])
    
    # Calculate total sales for the month
    total_sales = orders.aggregate(total=Sum('total_amount'))['total'] or 0

    # Calculate sales per product for the month
    order_items = OrderItem.objects.filter(order__in=orders)
    product_sales = order_items.values('product__name').annotate(
        total_quantity=Sum('quantity'),
        total_amount=Sum(F('quantity') * F('product__price'))
    ).order_by('-total_amount')

    # Calculate the unit price for each product
    for item in product_sales:
        item['unit_price'] = item['total_amount'] / item['total_quantity'] if item['total_quantity'] else 0

    if 'export' in request.GET:
        # Create an Excel workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = f"Sales Report {month}-{year}".replace("/", "-")

        # Write the headers
        headers = ['Product', 'Unit Price', 'Quantity Sold', 'Total Sales Amount']
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header

        # Write data to the Excel sheet
        for row_num, item in enumerate(product_sales, 2):
            ws[f"A{row_num}"] = item['product__name']
            ws[f"B{row_num}"] = item['unit_price']
            ws[f"C{row_num}"] = item['total_quantity']
            ws[f"D{row_num}"] = item['total_amount']

        # Add a row for total sales
        ws[f"A{row_num + 1}"] = 'Total Sales'
        ws[f"D{row_num + 1}"] = total_sales

        # Prepare the response to download the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Sales_Report_{month}_{year}.xlsx'
        wb.save(response)
        return response

    context = {
        'total_sales': total_sales,
        'month': f"{year}-{month:02d}",
        'product_sales': product_sales,
    }
    return render(request, 'monthly_sales.html', context)


from django.core.mail import send_mail
from django.shortcuts import render, redirect
from django.contrib import messages

@login_required(login_url=user_login)
def contact(request):
    if request.method == 'POST':
        name = request.POST['name']
        email = request.POST['email']
        subject = request.POST['subject']
        message = request.POST['message']
        
        # Prepare email message
        email_message = f"Name: {name}\nEmail: {email}\nSubject: {subject}\nMessage: {message}"

        try:
            send_mail(
                subject, 
                email_message, 
                email,  # from email
                ['aleeshasibi26@gmail.com'],  # to email (your host email)
            )
            messages.success(request, 'Your message has been sent successfully!')
        except Exception as e:
            messages.error(request, 'An error occurred. Please try again later.')
        
        return redirect('contact')
    
    return render(request, 'contact.html')


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import Order

def cancel_order(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    
    if order.status == 'processing' and order.user == request.user:
        order.status = 'canceled'
        order.save()
        messages.success(request, 'Your order has been canceled successfully.')
    else:
        messages.error(request, 'Order cannot be canceled.')
    
    return redirect('order_list_and_detail')